from django.core.management.base import BaseCommand
import os
import sys

class Command(BaseCommand):
    help = 'Export HTML email templates from Excel file'

    def handle(self, *args, **options):
        try:
            # تنظیم encoding برای خروجی
            sys.stdout.reconfigure(encoding='utf-8')
            
            # import مستقیم تابع بدون ایجاد Notification
            from ConfigurationChangeRequest.business import Notification
            
            # ایجاد نمونه ساده از Notification بدون وابستگی‌های پیچیده
            class SimpleNotification:
                def create_html_template_file_from_excel(self):
                    """
                    این تابع به ازای هر الگوی نامه یک فایل ایجاد می کند. 
                    این کار صرفا برای تست الگوهای نامه است
                    کافی است الگوهای جدید را از جدول مربوطه در دیتابیس اطلاع رسانی خروجی بگیریم
                    و در قالب فایل اکسل با نام زیر در مسیر زیر قرار دهیم
                    static\\ConfigurationChangeRequest\\email-template\\template.xlsx
                    """
                    try:
                        import openpyxl
                        from Config import settings
                        
                        # مسیر کامل فایل اکسل
                        excel_file_path = os.path.join(settings.BASE_DIR, 'static', 'ConfigurationChangeRequest', 'email-template', 'template.xlsx')
                        
                        # بررسی وجود فایل اکسل
                        if not os.path.exists(excel_file_path):
                            return {"success": False, "message": f"فایل اکسل در مسیر {excel_file_path} یافت نشد"}
                        
                        # باز کردن فایل اکسل
                        workbook = openpyxl.load_workbook(excel_file_path)
                        sheet = workbook['Sheet1']  # استفاده از اولین شیت
                        
                        # مسیر پوشه خروجی برای فایل‌های HTML
                        output_dir = os.path.join(settings.BASE_DIR, 'static', 'ConfigurationChangeRequest', 'email-template')
                        
                        # ایجاد پوشه در صورت عدم وجود
                        os.makedirs(output_dir, exist_ok=True)
                        
                        created_files = []
                        
                        # خواندن رکوردها از اکسل (فرض می‌کنیم ردیف اول شامل هدر است)
                        for row_num in range(2, sheet.max_row + 1):  # شروع از ردیف دوم
                            # خواندن مقدار سلول code (ستون اول)
                            code_cell = sheet.cell(row=row_num, column=1)
                            # خواندن مقدار سلول body (ستون پنجم)
                            body_cell = sheet.cell(row=row_num, column=5)
                            
                            # بررسی وجود مقادیر
                            if code_cell.value is None or body_cell.value is None:
                                continue
                            
                            code = str(code_cell.value).strip()
                            body_content = str(body_cell.value).strip()
                            
                            if not code or not body_content:
                                continue
                            
                            # نام فایل HTML
                            html_filename = f"{code}.html"
                            html_file_path = os.path.join(output_dir, html_filename)
                            
                            # محتوای HTML کامل
                            html_content = f"""<!DOCTYPE html>
<html lang="fa" dir="rtl">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{code}</title>
</head>
<body>
{body_content}
</body>
</html>"""
                            
                            # نوشتن فایل HTML
                            with open(html_file_path, 'w', encoding='utf-8') as html_file:
                                html_file.write(html_content)
                            
                            created_files.append(html_filename)
                        
                        # بستن فایل اکسل
                        workbook.close()
                        
                        return {
                            "success": True, 
                            "message": f"با موفقیت {len(created_files)} فایل HTML ایجاد شد",
                            "created_files": created_files
                        }
                        
                    except Exception as e:
                        return {"success": False, "message": f"خطا در ایجاد فایل‌های HTML: {str(e)}"}
            
            # ایجاد نمونه ساده
            notification = SimpleNotification()
            
            # فراخوانی تابع
            print("Starting HTML template export from Excel...")
            result = notification.create_html_template_file_from_excel()
            
            if result['success']:
                print(f"SUCCESS: {result['message']}")
                print(f"Exported files: {result['created_files']}")
                
                # نمایش فایل‌های ایجاد شده
                output_dir = os.path.join('static', 'ConfigurationChangeRequest', 'email-template')
                for filename in result['created_files']:
                    file_path = os.path.join(output_dir, filename)
                    if os.path.exists(file_path):
                        print(f"✓ {filename} exported successfully")
                    else:
                        print(f"✗ {filename} not found")
            else:
                print(f"ERROR: {result['message']}")
                
        except Exception as e:
            print(f"UNEXPECTED ERROR: {str(e)}")

    def handle(self, *args, **options):
        try:
            # تنظیم encoding برای خروجی
            sys.stdout.reconfigure(encoding='utf-8')
            
            # import مستقیم تابع بدون ایجاد Notification
            from ConfigurationChangeRequest.business import Notification
            
            # ایجاد نمونه ساده از Notification بدون وابستگی‌های پیچیده
            class SimpleNotification:
                def create_html_template_file_from_excel(self):
                    """
                    این تابع به ازای هر الگوی نامه یک فایل ایجاد می کند. 
                    این کار صرفا برای تست الگوهای نامه است
                    کافی است الگوهای جدید را از جدول مربوطه در دیتابیس اطلاع رسانی خروجی بگیریم
                    و در قالب فایل اکسل با نام زیر در مسیر زیر قرار دهیم
                    static\\ConfigurationChangeRequest\\email-template\\template.xlsx
                    """
                    try:
                        import openpyxl
                        from Config import settings
                        
                        # مسیر کامل فایل اکسل
                        excel_file_path = os.path.join(settings.BASE_DIR, 'static', 'ConfigurationChangeRequest', 'email-template', 'template.xlsx')
                        
                        # بررسی وجود فایل اکسل
                        if not os.path.exists(excel_file_path):
                            return {"success": False, "message": f"فایل اکسل در مسیر {excel_file_path} یافت نشد"}
                        
                        # باز کردن فایل اکسل
                        workbook = openpyxl.load_workbook(excel_file_path)
                        sheet = workbook['Sheet1']  # استفاده از اولین شیت
                        
                        # مسیر پوشه خروجی برای فایل‌های HTML
                        output_dir = os.path.join(settings.BASE_DIR, 'static', 'ConfigurationChangeRequest', 'email-template')
                        
                        # ایجاد پوشه در صورت عدم وجود
                        os.makedirs(output_dir, exist_ok=True)
                        
                        created_files = []
                        
                        # خواندن رکوردها از اکسل (فرض می‌کنیم ردیف اول شامل هدر است)
                        for row_num in range(2, sheet.max_row + 1):  # شروع از ردیف دوم
                            # خواندن مقدار سلول code (ستون اول)
                            code_cell = sheet.cell(row=row_num, column=1)
                            # خواندن مقدار سلول body (ستون پنجم)
                            body_cell = sheet.cell(row=row_num, column=5)
                            
                            # بررسی وجود مقادیر
                            if code_cell.value is None or body_cell.value is None:
                                continue
                            
                            code = str(code_cell.value).strip()
                            body_content = str(body_cell.value).strip()
                            
                            if not code or not body_content:
                                continue
                            
                            # نام فایل HTML
                            html_filename = f"{code}.html"
                            html_file_path = os.path.join(output_dir, html_filename)
                            
                            # محتوای HTML کامل
                            html_content = f"""<!DOCTYPE html>
                                                <html lang="fa" dir="rtl">
                                                <head>
                                                <meta charset="UTF-8">
                                                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                                                <title>{code}</title>
                                                </head>
                                                <body>
                                                {body_content}
                                                </body>
                                                </html>"""
                            
                            # نوشتن فایل HTML
                            with open(html_file_path, 'w', encoding='utf-8') as html_file:
                                html_file.write(html_content)
                            
                            created_files.append(html_filename)
                        
                        # بستن فایل اکسل
                        workbook.close()
                        
                        return {
                            "success": True, 
                            "message": f"با موفقیت {len(created_files)} فایل HTML ایجاد شد",
                            "created_files": created_files
                        }
                        
                    except Exception as e:
                        return {"success": False, "message": f"خطا در ایجاد فایل‌های HTML: {str(e)}"}
            
            # ایجاد نمونه ساده
            notification = SimpleNotification()
            
            # فراخوانی تابع
            print("Starting HTML template creation from Excel...")
            result = notification.create_html_template_file_from_excel()
            
            if result['success']:
                print(f"SUCCESS: {result['message']}")
                print(f"Created files: {result['created_files']}")
                
                # نمایش فایل‌های ایجاد شده
                output_dir = os.path.join('static', 'ConfigurationChangeRequest', 'email-template')
                for filename in result['created_files']:
                    file_path = os.path.join(output_dir, filename)
                    if os.path.exists(file_path):
                        print(f"✓ {filename} created successfully")
                    else:
                        print(f"✗ {filename} not found")
            else:
                print(f"ERROR: {result['message']}")
                
        except Exception as e:
            print(f"UNEXPECTED ERROR: {str(e)}")
