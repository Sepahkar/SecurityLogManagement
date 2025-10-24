from django.core.management.base import BaseCommand
import os
import sys
import re

class Command(BaseCommand):
    help = 'Update Excel file with HTML template files content'

    def handle(self, *args, **options):
        try:
            # تنظیم encoding برای خروجی
            sys.stdout.reconfigure(encoding='utf-8')
            
            # import مستقیم تابع بدون ایجاد Notification
            from ConfigurationChangeRequest.business import Notification
            
            # ایجاد نمونه ساده از Notification بدون وابستگی‌های پیچیده
            class SimpleNotification:
                def update_html_template_file_to_excel(self):
                    """
                    در صورتی که فایلهای الگوها را تغییر بدهیم و نیاز داشته باشیم که این تغییرات به دیتابیس منتقل شود
                    از این تابع استفاده می کنیم
                    این تابع الگوها را در فایل اکسل به روز می کند
                    فقط کافی است که اکسل مربوطه را در دیتابیس اطلاع رسانی بارگذاری کنیم
                    """
                    try:
                        import openpyxl
                        from Config import settings
                        
                        # مسیر کامل فایل اکسل
                        excel_file_path = os.path.join(settings.BASE_DIR, 'static', 'ConfigurationChangeRequest', 'email-template', 'template.xlsx')
                        
                        # بررسی وجود فایل اکسل
                        if not os.path.exists(excel_file_path):
                            return {"success": False, "message": f"فایل اکسل در مسیر {excel_file_path} یافت نشد"}
                        
                        # مسیر پوشه فایل‌های HTML
                        html_dir = os.path.join(settings.BASE_DIR, 'static', 'ConfigurationChangeRequest', 'email-template')
                        
                        # بررسی وجود پوشه HTML
                        if not os.path.exists(html_dir):
                            return {"success": False, "message": f"پوشه فایل‌های HTML در مسیر {html_dir} یافت نشد"}
                        
                        # باز کردن فایل اکسل
                        workbook = openpyxl.load_workbook(excel_file_path)
                        sheet = workbook['Sheet1']
                        
                        # ایجاد دیکشنری برای نگهداری کدها و ردیف‌های مربوطه
                        code_to_row = {}
                        
                        # خواندن کدهای موجود در اکسل
                        for row_num in range(2, sheet.max_row + 1):
                            code_cell = sheet.cell(row=row_num, column=1)
                            if code_cell.value is not None:
                                code = str(code_cell.value).strip()
                                if code:
                                    code_to_row[code] = row_num
                        
                        updated_files = []
                        new_records = []
                        
                        # خواندن فایل‌های HTML
                        html_files = [f for f in os.listdir(html_dir) if f.endswith('.html')]
                        
                        for html_file in html_files:
                            # استخراج کد از نام فایل
                            code = html_file.replace('.html', '')
                            
                            # خواندن محتوای فایل HTML
                            html_file_path = os.path.join(html_dir, html_file)
                            
                            try:
                                with open(html_file_path, 'r', encoding='utf-8') as f:
                                    html_content = f.read()
                                
                                # استخراج محتوای داخل تگ <body>
                                body_match = re.search(r'<body[^>]*>(.*?)</body>', html_content, re.DOTALL | re.IGNORECASE)
                                
                                if body_match:
                                    body_content = body_match.group(1).strip()
                                    
                                    # اگر رکورد با این کد وجود دارد، به‌روزرسانی کن
                                    if code in code_to_row:
                                        row_num = code_to_row[code]
                                        sheet.cell(row=row_num, column=5).value = body_content
                                        updated_files.append(code)
                                    else:
                                        # رکورد جدید ایجاد کن
                                        new_row = sheet.max_row + 1
                                        sheet.cell(row=new_row, column=1).value = code  # ستون Code
                                        sheet.cell(row=new_row, column=5).value = body_content  # ستون Body
                                        new_records.append(code)
                                        
                            except Exception as e:
                                print(f"خطا در خواندن فایل {html_file}: {str(e)}")
                                continue
                        
                        # ذخیره فایل اکسل
                        workbook.save(excel_file_path)
                        workbook.close()
                        
                        return {
                            "success": True,
                            "message": f"فایل اکسل با موفقیت به‌روزرسانی شد. {len(updated_files)} رکورد به‌روزرسانی و {len(new_records)} رکورد جدید اضافه شد",
                            "updated_files": updated_files,
                            "new_records": new_records
                        }
                        
                    except Exception as e:
                        return {"success": False, "message": f"خطا در به‌روزرسانی فایل اکسل: {str(e)}"}
            
            # ایجاد نمونه ساده
            notification = SimpleNotification()
            
            # فراخوانی تابع
            print("Starting Excel file update from HTML templates...")
            result = notification.update_html_template_file_to_excel()
            
            if result['success']:
                print(f"SUCCESS: {result['message']}")
                
                if result['updated_files']:
                    print(f"Updated files: {result['updated_files']}")
                
                if result['new_records']:
                    print(f"New records: {result['new_records']}")
                    
            else:
                print(f"ERROR: {result['message']}")
                
        except Exception as e:
            print(f"UNEXPECTED ERROR: {str(e)}")
